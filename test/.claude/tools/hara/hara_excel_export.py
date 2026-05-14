#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit("缺少 openpyxl，请先安装：pip install openpyxl") from exc

try:
    from hara_schema_columns import SHEET_COLUMNS, EXCEL_DISPLAY_HEADERS, is_nan_like
except ImportError:  # pragma: no cover
    from .hara_schema_columns import SHEET_COLUMNS, EXCEL_DISPLAY_HEADERS, is_nan_like

SHEET_KEY_ALIASES = {
    "DeriveMF": ["DeriveMF", "derive_mf"],
    "MF and Vehicle Hazards": ["MF and Vehicle Hazards", "mf_vehicle_hazards"],
    "HARA": ["HARA", "hara"],
    "SG_Sum": ["SG_Sum", "sg_sum"],
}

def load_json(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8-sig").strip()
    if text.startswith("```json"):
        text = text[len("```json"):].strip()
    if text.startswith("```"):
        text = text[3:].strip()
    if text.endswith("```"):
        text = text[:-3].strip()
    data = json.loads(text)
    if not isinstance(data, dict):
        raise SystemExit("Top-level JSON must be an object")
    return data

def as_rows(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    return []

def get_rows(data: dict[str, Any], sheet_name: str) -> list[dict[str, Any]]:
    for key in SHEET_KEY_ALIASES.get(sheet_name, [sheet_name]):
        if key in data:
            return as_rows(data[key])
    return []

def as_text(value: Any) -> Any:
    if is_nan_like(value):
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value

def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="1F2937")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        values = [str(c.value or "") for c in column_cells[:100]]
        max_len = max([len(v) for v in values] + [8])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 45)
    ws.row_dimensions[1].height = 32

def write_table(wb: Workbook, sheet_name: str, columns: list[str], rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(sheet_name)
    display_headers = [EXCEL_DISPLAY_HEADERS.get(h, h) for h in columns]
    ws.append(display_headers)
    for row in rows:
        ws.append([as_text(row.get(h, "nan")) for h in columns])
    style_sheet(ws)

def write_warnings(wb: Workbook, warnings: list[Any]) -> None:
    if not warnings:
        return
    ws = wb.create_sheet("Validation_Warnings")
    headers = ["level", "stage", "sheet", "row", "field", "value", "message"]
    ws.append(headers)
    for item in warnings:
        if isinstance(item, dict):
            ws.append([as_text(item.get(h, "nan")) for h in headers])
        else:
            ws.append(["WARNING", "unknown", "unknown", "", "", "", as_text(item)])
    style_sheet(ws)
def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    data = load_json(Path(args.input))
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, columns in SHEET_COLUMNS.items():
        write_table(wb, sheet_name, columns, get_rows(data, sheet_name))

    warnings = data.get("Validation_Warnings") or data.get("validation_warnings") or []
    write_warnings(wb, warnings if isinstance(warnings, list) else [])

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"[OK] Excel written: {out}")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
